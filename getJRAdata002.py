# 新規でExeclを作るときの処理を作る
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

if __name__ == '__main__':

    column_a = ['年月日', ' 開催場', '回', '日', 'レース', '発走時刻', 'レース名', '競争条件', '距離', '登録頭数',
                '芝／ダート', '1着馬番', '2着馬番', '3着馬番', '4着馬番', '5着馬番', '6着馬番', '同着用予備',
                '同着用予備.1', '同着用予備.2', '同着用予備.3', '同着用予備.4', '単勝1', '単勝2', '複勝1', '複勝2',
                '複勝3', '複勝4', '枠1', '枠1.1', '枠連配当1', '枠2', '枠2.1', '枠連配当2', '馬1',
                '馬1.1', '馬連配当1', '馬2', '馬2.1', '馬連配当2', 'ワイド1', 'ワイド1.1', 'ワイド配当1',
                'ワイド2', 'ワイド2.1', 'ワイド配当2', 'ワイド3', 'ワイド3.1', 'ワイド配当3', 'ワイド4',
                'ワイド4.1', 'ワイド配当4', 'ワイド5', 'ワイド5.1', 'ワイド配当5', '馬単1', '馬単1.1',
                '馬単配当1', '馬単2', '馬単2.1', '馬単配当2', '3連複1', '3連複1.1', '3連複1.2',
                '3連複配当1', '3連複2', '3連複2.1', '3連複2.2', '3連複配当2', '3連単1', '3連単1.1',
                '3連単1.2', '3連単配当1', '3連単2', '3連単2.1', '3連単2.2', '3連単配当2']

    column_b = ['年月日', ' 開催場', '回', '日', 'レース', '競争条件', '距離', '登録頭数', '発走時刻', '馬番',
                '枠番', '馬名', '騎手', '斤量', '減量', '調教師', '種牡馬', '母名', '馬主', '母父', '人気',
                'オッズ', '着順', '単勝', '複勝']

    column_c = ['年月日', '開催', '発走時刻', 'レースナンバー', '競争条件', '距離', '芝／ダート', '馬名', '馬主名',
                '調教師名', '父', '母', '母の父', '騎手']

    df_a = pd.DataFrame(columns=column_a)
    list = []
    for col in df_a.columns:
        list.append("test")
    sr_a = pd.Series(list, index=df_a.columns)
    df_a = df_a.append(sr_a, ignore_index=True)

    df_b = pd.DataFrame(columns=column_b)
    list = []
    for col in df_b.columns:
        list.append("test")
    sr_b = pd.Series(list, index=df_b.columns)
    df_b = df_b.append(sr_b, ignore_index=True)

    df_c = pd.DataFrame(columns=column_c)
    list = []
    for col in df_c.columns:
        list.append("test")
    sr_c = pd.Series(list, index=df_c.columns)
    df_c = df_c.append(sr_c, ignore_index=True)

    # Excelワークブックの生成
    wb = Workbook()
    ws_a = wb.active

    ws_a.title = "A"
    rows = dataframe_to_rows(df_a, index=False, header=True)  # openpyxlのユーティリティを使用

    # ワークシートへデータを書き込む
    # row_start_idx = 0
    # col_start_idx = 0
    # for row_no, row in enumerate(rows, row_start_idx):
    #     for col_no, value in enumerate(row, col_start_idx):
    #         ws.cell(row=row_no + 1, column=col_no + 1, value=value)  # 1セルづつ書込む

    for i, row in enumerate(dataframe_to_rows(df_a, index=False, header=True)):
        ws_a.append(row)

    ws_b = wb.create_sheet(title="B")
    for i, row in enumerate(dataframe_to_rows(df_b, index=False, header=True)):
        ws_b.append(row)

    ws_c = wb.create_sheet(title="C")
    for i, row in enumerate(dataframe_to_rows(df_c, index=False, header=True)):
        ws_c.append(row)

    date_from = datetime(2021, 1, 1)
    date_to = datetime(2021, 12, 31)
    wb.save("JRAdata_{}_{}.xlsx".format(
        date_from.strftime("%Y%m%d"),
        date_to.strftime("%Y%m%d")))
