from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

if __name__ == '__main__':
    df_data = pd.DataFrame(
        {'販売店': ['伊勢店', '伊勢店', '伊勢店', '伊勢店 小計', '出雲店', '出雲店', '出雲店', '出雲店 小計', '靖国店', '靖国店', '靖国店', '靖国店 小計', '総計'],
         '商品': ['おふだ', 'おみくじ', 'お守り', '', 'おふだ', 'おみくじ', 'お守り', '', 'おふだ', 'おみくじ', 'お守り', '', ''],
         '1月1日': [6321, 176, 1140, 7637, 6018, 273, 101, 6392, 2847, 2203, 3939, 8989, 23018],
         '1月2日': [323, 27, 507, 857, 297, 4, 101, 402, 2598, 1874, 2736, 7208, 8467],
         '1月3日': [192, 78, 341, 611, 48, 81, 84, 213, 2341, 1049, 3092, 6482, 7306]})

    # Excelワークブックの生成
    wb = Workbook()
    ws = wb.active

    ws.title = '2020年'
    rows = dataframe_to_rows(df_data, index=False, header=True)  # openpyxlのユーティリティを使用

    # ワークシートへデータを書き込む
    row_start_idx = 3
    col_start_idx = 2
    for row_no, row in enumerate(rows, row_start_idx):
        for col_no, value in enumerate(row, col_start_idx):
            ws.cell(row=row_no, column=col_no, value=value)  # 1セルづつ書込む

    ws.cell(row=1, column=1, value='■Excel出力のサンプル')  # タイトル

    wb.save('sample.xlsx')
