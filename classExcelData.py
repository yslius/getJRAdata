# シートA、シートBの情報取得試してみる
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import glob


class ExcelData():
    column_a = ['年月日', ' 開催場', '回', '日', 'レース', '発走時刻', 'レース名', '競争条件', '距離', '登録頭数',
                '芝／ダート', '1着馬番', '2着馬番', '3着馬番', '4着馬番', '5着馬番', '6着馬番', '同着用予備',
                '同着用予備.1', '同着用予備.2', '同着用予備.3', '同着用予備.4', '単勝1', '単勝2', '複勝1', '複勝2',
                '複勝3', '複勝4', '枠1', '枠1.1', '枠連配当1', '枠2', '枠2.1', '枠連配当2', '馬1',
                '馬1.1', '馬連配当1', '馬2', '馬2.1', '馬連配当2', 'ワイド1', 'ワイド1.1', 'ワイド配当1',
                'ワイド2', 'ワイド2.1', 'ワイド配当2', 'ワイド3', 'ワイド3.1', 'ワイド配当3', 'ワイド4',
                'ワイド4.1', 'ワイド配当4', 'ワイド5', 'ワイド5.1', 'ワイド配当5', '馬単1', '馬単1.1',
                '馬単配当1', '馬単2', '馬単2.1', '馬単配当2', '3連複1', '3連複1.1', '3連複1.2',
                '3連複配当1', '3連複2', '3連複2.1', '3連複2.2', '3連複配当2', '3連単1', '3連単1.1',
                '3連単1.2', '3連単配当1', '3連単2', '3連単2.1', '3連単2.2', '3連単配当2']

    column_b = ['年月日', ' 開催場', '回', '日', 'レース', '競争条件', '距離', '登録頭数', '発走時刻', '馬番',
                '枠番', '馬名', '騎手', '斤量', '減量', '調教師', '種牡馬', '母名', '馬主', '母父', '人気',
                'オッズ', '着順', '単勝', '複勝']

    column_c = ['年月日', '開催', '発走時刻', 'レースナンバー', '競争条件', '距離', '芝／ダート', '馬名', '馬主名',
                '調教師名', '父', '母', '母の父', '騎手']
    name_excel = "JRAdata"
    wb = ""

    def __init__(self):
        self.df_a = pd.DataFrame(columns=self.column_a)
        self.df_b = pd.DataFrame(columns=self.column_b)
        self.df_c = pd.DataFrame(columns=self.column_c)
        self.list_a = []
        self.list_b = []
        self.list_c = []

    def append_list(self, select, data):
        """
        リストにデータを追加する
        """
        if select == "A":
            self.list_a.append(data)
        elif select == "B":
            self.list_b.append(data)
        else:
            self.list_c.append(data)
        # リストの長さチェックする
        self.check_list(select)

    def check_list(self, select):
        """
        リストの長さをチェックして、カラム数と一致してたらdfに追加する
        """
        if select == "A":
            if len(self.list_a) == len(self.df_a.columns):
                self.append_df(select)
                self.list_a = []
        elif select == "B":
            if len(self.list_b) == len(self.df_b.columns):
                self.append_df(select)
                self.list_b = []
        else:
            if len(self.list_c) == len(self.df_c.columns):
                self.append_df(select)
                self.list_c = []

    def append_df(self, select):
        """
        dfにリストを追加する
        """
        if select == "A":
            sr_a = pd.Series(self.list_a, index=self.df_a.columns)
            self.df_a = self.df_a.append(sr_a, ignore_index=True)
        elif select == "B":
            sr_b = pd.Series(self.list_b, index=self.df_b.columns)
            self.df_b = self.df_b.append(sr_b, ignore_index=True)
        else:
            sr_c = pd.Series(self.list_c, index=self.df_c.columns)
            self.df_c = self.df_c.append(sr_c, ignore_index=True)


    def read_excel(self):
        files = glob.glob("./{}*.xlsx".format(self.name_excel))
        print("files:{}".format(files))
        if len(files) == 0:
            print("Excelが見つかりません")
            return False
        if len(files) > 1:
            print("Excelが複数あります")
            return False

        # Excelワークブックの読み込み
        self.wb = openpyxl.load_workbook(files[0])
        self.df_c = pd.read_excel(files[0], sheet_name="C")

        return True

    def create_new_excel(self, date_from, date_to):
        """
        新しいExcelを作成してシートA、Bを完成させる
        """
        self.wb = openpyxl.Workbook()
        ws_a = self.wb.active

        ws_a.title = "A"
        for i, row in enumerate(dataframe_to_rows(self.df_a, index=False, header=True)):
            ws_a.append(row)

        ws_b = self.wb.create_sheet(title="B")
        for i, row in enumerate(dataframe_to_rows(self.df_b, index=False, header=True)):
            ws_b.append(row)

        ws_c = self.wb.create_sheet(title="C")
        for i, row in enumerate(dataframe_to_rows(self.df_c, index=False, header=True)):
            ws_c.append(row)

        self.wb.save("{}_{}_{}.xlsx".format(
            self.name_excel,
            date_from.strftime("%Y%m%d"),
            date_to.strftime("%Y%m%d")))

