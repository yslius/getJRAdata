# Execlを読み取る処理を作る
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

if __name__ == '__main__':
    df_a = pd.read_excel("JRAdata.xlsx", sheet_name="A")
    print(df_a.columns)
    df_b = pd.read_excel("JRAdata.xlsx", sheet_name="B")
    print(df_b.columns)
    df_c = pd.read_excel("JRAdata.xlsx", sheet_name="C")
    print(df_c.columns)

    # Excelワークブックの読み込み
    wb = openpyxl.load_workbook("JRAdata.xlsx")

    ws_b = wb["B"]

    # データ読み取り
    df = pd.DataFrame(ws_b.values)
    print(df.columns)

    # 一旦削除して追加してみる
    df.drop(range(len(df)), inplace=True)
    list = []
    for col in df.columns:
        list.append("test")
    sr = pd.Series(list, index=df.columns)
    df = df.append(sr, ignore_index=True)

    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet(title="A")

    # 一旦全部消してから書き込む
    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.value = None
    # for i, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
    #     ws.append(row)

    # 追加して書き込む
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        ws_b.append(row)

    wb.save("JRAdata_.xlsx")

    # df = pd.read_excel("JRAdata.xlsx", sheet_name="B")
    # print(df)
