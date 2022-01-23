# シートA、シートBの情報取得試してみる
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import glob


class ScrapeJRA():
    name_logfile = "getJRAdata"
    url_base = "https://www.jra.go.jp/JRADB/accessS.html"
    headers = {"User-Agent": "Mozilla/5.0"}
    session = ""
    req = ""
    obj_param = {}

    def __init__(self):
        self.set_obj_param()
        self.session = requests.session()

    def print_write_log(self, str_disp):
        print_log = "{} {}".format(
            datetime.now().strftime("%H:%M:%S"), str_disp)
        print(print_log)
        path_logfile = "log/{}_{}.log".format(
            self.name_logfile, datetime.now().strftime("%Y%m%d"))
        with open(path_logfile, "a+") as file_object:
            file_object.writelines(print_log + "\n")

    def next_page(self, payload):
        print("next_page Start")

        self.req = self.session.post(self.url_base, headers=self.headers, data=payload)
        print("status_code:".format(self.req.status_code))
        if self.req.status_code != 200:
            print("status_code != 200")
            return False
        self.req.encoding = self.req.apparent_encoding

        print("next_page End")

        return True

    def next_page_date(self, year, month):
        """
        指定の日付のページにいく
        """
        print("next_page_place Start")

        cname = self.obj_param[(year + month)[2:]]
        payload = {"cname": "pw01skl10" + year + month + "/" + cname}
        print("payload:{}".format(payload))

        if not self.next_page(payload):
            return False
        if "パラメータエラー" in self.req.text:
            print("パラメータエラー")
            return False

        print("next_page_place End")

        return True

    def next_page_place(self):
        """
        指定の日付の会場ごとのページにいく
        """
        print("next_page_place Start")

        soup = BeautifulSoup(self.req.text, "lxml")
        ele_div = soup.select("div.link_list")
        ele_a = ele_div[0].select("a")
        cname = self.get_cname(ele_a[0].attrs["onclick"])
        payload = {"cname": cname}
        print("payload:{}".format(payload))

        if not self.next_page(payload):
            return False
        if "パラメータエラー" in self.req.text:
            print("パラメータエラー")
            return False

        print("next_page_place End")

        return True

    def next_page_place_race(self):
        """
        レース詳細ページにいく
        """
        print("next_page_place_race Start")

        soup = BeautifulSoup(self.req.text, 'lxml')
        ele_th = soup.select("th.race_num")
        ele_a = ele_th[1].select_one("a")
        cname = self.get_cname(ele_a.attrs["onclick"])
        payload = {"cname": cname}
        print("payload:{}".format(payload))

        if not self.next_page(payload):
            return False
        if "パラメータエラー" in self.req.text:
            print("パラメータエラー")
            return False

        print("next_page_place_race End")

        return True

    def next_page_umadata(self):
        """
        馬ごとの情報を見る
        """
        print("next_page_umadata Start")

        soup = BeautifulSoup(self.req.text, 'lxml')
        ele_td = soup.select("td.horse")
        ele_a = ele_td[0].select_one("a")
        cname = self.get_cname(ele_a.attrs["onclick"])
        payload = {"cname": cname}
        print("payload:{}".format(payload))

        if not self.next_page(payload):
            return False
        if "パラメータエラー" in self.req.text:
            print("パラメータエラー")
            return False

        print("next_page_umadata End")

        return True

    def get_cname(self, str_doaction):
        tmp_array = str_doaction.split(",")
        tmp = tmp_array[1]

        return tmp[2:-3]

    def set_obj_param(self):
        self.obj_param["2201"] = "B4"
        self.obj_param["2202"] = "82"
        self.obj_param["2203"] = "50"
        self.obj_param["2204"] = "1E"
        self.obj_param["2205"] = "EC"
        self.obj_param["2206"] = "BA"
        self.obj_param["2207"] = "88"
        self.obj_param["2208"] = "56"
        self.obj_param["2209"] = "24"
        self.obj_param["2210"] = "93"
        self.obj_param["2211"] = "61"
        self.obj_param["2212"] = "2F"
        self.obj_param["2101"] = "0F"
        self.obj_param["2102"] = "DD"
        self.obj_param["2103"] = "AB"
        self.obj_param["2104"] = "92"
        self.obj_param["2105"] = "60"
        self.obj_param["2106"] = "2E"
        self.obj_param["2107"] = "FC"
        self.obj_param["2108"] = "CA"
        self.obj_param["2109"] = "98"
        self.obj_param["2110"] = "07"
        self.obj_param["2111"] = "D5"
        self.obj_param["2112"] = "A3"
        self.obj_param["2001"] = "83"
        self.obj_param["2002"] = "51"
        self.obj_param["2003"] = "1F"
        self.obj_param["2004"] = "ED"
        self.obj_param["2005"] = "BB"
        self.obj_param["2006"] = "89"
        self.obj_param["2007"] = "57"
        self.obj_param["2008"] = "25"
        self.obj_param["2009"] = "F3"
        self.obj_param["2010"] = "62"
        self.obj_param["2011"] = "30"
        self.obj_param["2012"] = "FE"
        self.obj_param["1901"] = "04"
        self.obj_param["1902"] = "D2"
        self.obj_param["1903"] = "A0"
        self.obj_param["1904"] = "6E"
        self.obj_param["1905"] = "3C"
        self.obj_param["1906"] = "0A"
        self.obj_param["1907"] = "D8"
        self.obj_param["1908"] = "A6"
        self.obj_param["1909"] = "74"
        self.obj_param["1910"] = "E3"
        self.obj_param["1911"] = "B1"
        self.obj_param["1912"] = "7F"
        self.obj_param["1801"] = "78"
        self.obj_param["1802"] = "46"
        self.obj_param["1803"] = "14"
        self.obj_param["1804"] = "E2"
        self.obj_param["1805"] = "B0"
        self.obj_param["1806"] = "7E"
        self.obj_param["1807"] = "4C"
        self.obj_param["1808"] = "1A"
        self.obj_param["1809"] = "E8"
        self.obj_param["1810"] = "57"
        self.obj_param["1811"] = "25"
        self.obj_param["1812"] = "F3"
        self.obj_param["1701"] = "EC"
        self.obj_param["1702"] = "BA"
        self.obj_param["1703"] = "88"
        self.obj_param["1704"] = "56"
        self.obj_param["1705"] = "24"
        self.obj_param["1706"] = "F2"
        self.obj_param["1707"] = "C0"
        self.obj_param["1708"] = "8E"
        self.obj_param["1709"] = "5C"
        self.obj_param["1710"] = "CB"
        self.obj_param["1711"] = "99"
        self.obj_param["1712"] = "67"
        self.obj_param["1601"] = "60"
        self.obj_param["1602"] = "2E"
        self.obj_param["1603"] = "FC"
        self.obj_param["1604"] = "CA"
        self.obj_param["1605"] = "98"
        self.obj_param["1606"] = "66"
        self.obj_param["1607"] = "34"
        self.obj_param["1608"] = "02"
        self.obj_param["1609"] = "D0"
        self.obj_param["1610"] = "3F"
        self.obj_param["1611"] = "0D"
        self.obj_param["1612"] = "DB"
        self.obj_param["1501"] = "D4"
        self.obj_param["1502"] = "A2"
        self.obj_param["1503"] = "70"
        self.obj_param["1504"] = "3E"
        self.obj_param["1505"] = "0C"
        self.obj_param["1506"] = "DA"
        self.obj_param["1507"] = "A8"
        self.obj_param["1508"] = "76"
        self.obj_param["1509"] = "44"
        self.obj_param["1510"] = "B3"
        self.obj_param["1511"] = "81"
        self.obj_param["1512"] = "4F"
        self.obj_param["1401"] = "48"
        self.obj_param["1402"] = "16"
        self.obj_param["1403"] = "E4"
        self.obj_param["1404"] = "B2"
        self.obj_param["1405"] = "80"
        self.obj_param["1406"] = "4E"
        self.obj_param["1407"] = "1C"
        self.obj_param["1408"] = "EA"
        self.obj_param["1409"] = "B8"
        self.obj_param["1410"] = "27"
        self.obj_param["1411"] = "F5"
        self.obj_param["1412"] = "C3"
        self.obj_param["1301"] = "BC"
        self.obj_param["1302"] = "8A"
        self.obj_param["1303"] = "58"
        self.obj_param["1304"] = "26"
        self.obj_param["1305"] = "F4"
        self.obj_param["1306"] = "C2"
        self.obj_param["1307"] = "90"
        self.obj_param["1308"] = "5E"
        self.obj_param["1309"] = "2C"
        self.obj_param["1310"] = "9B"
        self.obj_param["1311"] = "69"
        self.obj_param["1312"] = "37"
        self.obj_param["1201"] = "30"
        self.obj_param["1202"] = "FE"
        self.obj_param["1203"] = "CC"
        self.obj_param["1204"] = "9A"
        self.obj_param["1205"] = "68"
        self.obj_param["1206"] = "36"
        self.obj_param["1207"] = "04"
        self.obj_param["1208"] = "D2"
        self.obj_param["1209"] = "A0"
        self.obj_param["1210"] = "0F"
        self.obj_param["1211"] = "DD"
        self.obj_param["1212"] = "AB"
        self.obj_param["1101"] = "A4"
        self.obj_param["1102"] = "72"
        self.obj_param["1103"] = "40"
        self.obj_param["1104"] = "0E"
        self.obj_param["1105"] = "DC"
        self.obj_param["1106"] = "AA"
        self.obj_param["1107"] = "78"
        self.obj_param["1108"] = "46"
        self.obj_param["1109"] = "14"
        self.obj_param["1110"] = "83"
        self.obj_param["1111"] = "51"
        self.obj_param["1112"] = "1F"
        self.obj_param["1001"] = "18"
        self.obj_param["1002"] = "E6"
        self.obj_param["1003"] = "B4"
        self.obj_param["1004"] = "82"
        self.obj_param["1005"] = "50"
        self.obj_param["1006"] = "1E"
        self.obj_param["1007"] = "EC"
        self.obj_param["1008"] = "BA"
        self.obj_param["1009"] = "88"
        self.obj_param["1010"] = "F7"
        self.obj_param["1011"] = "C5"
        self.obj_param["1012"] = "93"
        self.obj_param["0901"] = "99"
        self.obj_param["0902"] = "67"
        self.obj_param["0903"] = "35"
        self.obj_param["0904"] = "03"
        self.obj_param["0905"] = "D1"
        self.obj_param["0906"] = "9F"
        self.obj_param["0907"] = "6D"
        self.obj_param["0908"] = "3B"
        self.obj_param["0909"] = "09"
        self.obj_param["0910"] = "78"
        self.obj_param["0911"] = "46"
        self.obj_param["0912"] = "14"
        self.obj_param["0801"] = "0D"
        self.obj_param["0802"] = "DB"
        self.obj_param["0803"] = "A9"
        self.obj_param["0804"] = "77"
        self.obj_param["0805"] = "45"
        self.obj_param["0806"] = "13"
        self.obj_param["0807"] = "E1"
        self.obj_param["0808"] = "AF"
        self.obj_param["0809"] = "7D"
        self.obj_param["0810"] = "EC"
        self.obj_param["0811"] = "BA"
        self.obj_param["0812"] = "88"
        self.obj_param["0701"] = "81"
        self.obj_param["0702"] = "4F"
        self.obj_param["0703"] = "1D"
        self.obj_param["0704"] = "EB"
        self.obj_param["0705"] = "B9"
        self.obj_param["0706"] = "87"
        self.obj_param["0707"] = "55"
        self.obj_param["0708"] = "23"
        self.obj_param["0709"] = "F1"
        self.obj_param["0710"] = "60"
        self.obj_param["0711"] = "2E"
        self.obj_param["0712"] = "FC"
        self.obj_param["0601"] = "F5"
        self.obj_param["0602"] = "C3"
        self.obj_param["0603"] = "91"
        self.obj_param["0604"] = "5F"
        self.obj_param["0605"] = "2D"
        self.obj_param["0606"] = "FB"
        self.obj_param["0607"] = "C9"
        self.obj_param["0608"] = "97"
        self.obj_param["0609"] = "65"
        self.obj_param["0610"] = "D4"
        self.obj_param["0611"] = "A2"
        self.obj_param["0612"] = "70"
        self.obj_param["0501"] = "69"
        self.obj_param["0502"] = "37"
        self.obj_param["0503"] = "05"
        self.obj_param["0504"] = "D3"
        self.obj_param["0505"] = "A1"
        self.obj_param["0506"] = "6F"
        self.obj_param["0507"] = "3D"
        self.obj_param["0508"] = "0B"
        self.obj_param["0509"] = "D9"
        self.obj_param["0510"] = "48"
        self.obj_param["0511"] = "16"
        self.obj_param["0512"] = "E4"
        self.obj_param["0401"] = "DD"
        self.obj_param["0402"] = "AB"
        self.obj_param["0403"] = "79"
        self.obj_param["0404"] = "47"
        self.obj_param["0405"] = "15"
        self.obj_param["0406"] = "E3"
        self.obj_param["0407"] = "B1"
        self.obj_param["0408"] = "7F"
        self.obj_param["0409"] = "4D"
        self.obj_param["0410"] = "BC"
        self.obj_param["0411"] = "8A"
        self.obj_param["0412"] = "58"
        self.obj_param["0301"] = "51"
        self.obj_param["0302"] = "1F"
        self.obj_param["0303"] = "ED"
        self.obj_param["0304"] = "BB"
        self.obj_param["0305"] = "89"
        self.obj_param["0306"] = "57"
        self.obj_param["0307"] = "25"
        self.obj_param["0308"] = "F3"
        self.obj_param["0309"] = "C1"
        self.obj_param["0310"] = "30"
        self.obj_param["0311"] = "FE"
        self.obj_param["0312"] = "CC"
        self.obj_param["0201"] = "C5"
        self.obj_param["0202"] = "93"
        self.obj_param["0203"] = "61"
        self.obj_param["0204"] = "2F"
        self.obj_param["0205"] = "FD"
        self.obj_param["0206"] = "CB"
        self.obj_param["0207"] = "99"
        self.obj_param["0208"] = "67"
        self.obj_param["0209"] = "35"
        self.obj_param["0210"] = "A4"
        self.obj_param["0211"] = "72"
        self.obj_param["0212"] = "40"
        self.obj_param["0101"] = "39"
        self.obj_param["0102"] = "07"
        self.obj_param["0103"] = "D5"
        self.obj_param["0104"] = "A3"
        self.obj_param["0105"] = "71"
        self.obj_param["0106"] = "3F"
        self.obj_param["0107"] = "0D"
        self.obj_param["0108"] = "DB"
        self.obj_param["0109"] = "A9"
        self.obj_param["0110"] = "18"
        self.obj_param["0111"] = "E6"
        self.obj_param["0112"] = "B4"
        self.obj_param["0001"] = "AD"
        self.obj_param["0002"] = "7B"
        self.obj_param["0003"] = "49"
        self.obj_param["0004"] = "17"
        self.obj_param["0005"] = "E5"
        self.obj_param["0006"] = "B3"
        self.obj_param["0007"] = "81"
        self.obj_param["0008"] = "4F"
        self.obj_param["0009"] = "1D"
        self.obj_param["0010"] = "8C"
        self.obj_param["0011"] = "5A"
        self.obj_param["0012"] = "28"



class ExcelData():
    column_a = ['年月日', ' 開催場', '回', '日', 'レース', '発走時刻', 'レース名', '競争条件', '距離', '登録頭数',
                '芝／ダート', '1着馬番', '2着馬番', '3着馬番', '4着馬番', '5着馬番', '6着馬番', '同着用予備',
                '同着用予備.1', '同着用予備.2', '同着用予備.3', '同着用予備.4', '単勝1', '単勝2', '複勝1', '複勝2',
                '複勝3', '複勝4', '枠1', '枠1.1', '枠連配当1', '枠2', '枠2.1', '枠連配当2', '馬1',
                '馬1.1', '馬連配当1', '馬2', '馬2.1', '馬連配当2', 'ワイド1', 'ワイド1.1', 'ワイド配当1',
                'ワイド2', 'ワイド2.1', 'ワイド配当2', 'ワイド3', 'ワイド3.1', 'ワイド配当3', 'ワイド4',
                'ワイド4.1', 'ワイド配当4', 'ワイド5', 'ワイド5.1', 'ワイド配当5', '馬単1', '馬単1.1',
                '馬単配当1', '馬単2', '馬単2.1', '馬単配当2', '3連複1', '3連複1.1', '3連複1.2',
                '3連複配当1', '3連複2', '3連複2.1', '3連複2.2', '3連複配当2', '3連単1', '3連単1.1',
                '3連単1.2', '3連単配当1', '3連単2', '3連単2.1', '3連単2.2', '3連単配当2']

    column_b = ['年月日', ' 開催場', '回', '日', 'レース', '競争条件', '距離', '登録頭数', '発走時刻', '馬番',
                '枠番', '馬名', '騎手', '斤量', '減量', '調教師', '種牡馬', '母名', '馬主', '母父', '人気',
                'オッズ', '着順', '単勝', '複勝']

    column_c = ['年月日', '開催', '発走時刻', 'レースナンバー', '競争条件', '距離', '芝／ダート', '馬名', '馬主名',
                '調教師名', '父', '母', '母の父', '騎手']
    name_excel = "JRAdata"
    wb = ""

    def __init__(self):
        self.df_a = pd.DataFrame(columns=self.column_a)
        self.df_b = pd.DataFrame(columns=self.column_b)
        self.df_c = pd.DataFrame(columns=self.column_c)
        self.list_a = []
        self.list_b = []
        self.list_c = []

    def append_list(self, select, data):
        """
        リストにデータを追加する
        """
        if select == "A":
            self.list_a.append(data)
        elif select == "B":
            self.list_b.append(data)
        else:
            self.list_c.append(data)
        # リストの長さチェックする
        self.check_list(select)

    def check_list(self, select):
        """
        リストの長さをチェックして、カラム数と一致してたらdfに追加する
        """
        if select == "A":
            if len(self.list_a) == len(self.df_a.columns):
                self.append_df(select)
                self.list_a = []
        elif select == "B":
            if len(self.list_b) == len(self.df_b.columns):
                self.append_df(select)
                self.list_b = []
        else:
            if len(self.list_c) == len(self.df_c.columns):
                self.append_df(select)
                self.list_c = []

    def append_df(self, select):
        """
        dfにリストを追加する
        """
        if select == "A":
            sr_a = pd.Series(self.list_a, index=self.df_a.columns)
            self.df_a = self.df_a.append(sr_a, ignore_index=True)
        elif select == "B":
            sr_b = pd.Series(self.list_b, index=self.df_b.columns)
            self.df_b = self.df_b.append(sr_b, ignore_index=True)
        else:
            sr_c = pd.Series(self.list_c, index=self.df_c.columns)
            self.df_c = self.df_c.append(sr_c, ignore_index=True)


    def read_excel(self):
        files = glob.glob("./{}*.xlsx".format(self.name_excel))
        print("files:{}".format(files))
        if len(files) == 0 or len(files) > 1:
            print("Excelが見つかりません")
            return False
        if len(files) > 1:
            print("Excelが複数あります")
            return False

        # Excelワークブックの読み込み
        self.wb = openpyxl.load_workbook(files[0])
        self.df_c = pd.read_excel(files[0], sheet_name="C")

        return True

    def create_new_excel(self, date_from, date_to):
        """
        新しいExcelを作成してシートA、Bを完成させる
        """
        self.wb = openpyxl.Workbook()
        ws_a = self.wb.active

        ws_a.title = "A"
        for i, row in enumerate(dataframe_to_rows(self.df_a, index=False, header=True)):
            ws_a.append(row)

        ws_b = self.wb.create_sheet(title="B")
        for i, row in enumerate(dataframe_to_rows(self.df_b, index=False, header=True)):
            ws_b.append(row)

        ws_c = self.wb.create_sheet(title="C")
        for i, row in enumerate(dataframe_to_rows(self.df_c, index=False, header=True)):
            ws_c.append(row)

        self.wb.save("{}_{}_{}.xlsx".format(
            self.name_excel,
            date_from.strftime("%Y%m%d"),
            date_to.strftime("%Y%m%d")))


def get_data_newexcel(date_from, date_to):
    # インスタンスを作る
    exceldata = ExcelData()
    scrapejra = ScrapeJRA()

    # レース結果を選択する
    scrapejra.next_page({"cname": "pw01sli00/AF"})

    # 過去レース結果検索を選択する
    scrapejra.next_page({"cname": "pw01skl00999999/B3"})

    # 例えば2020年6月を選択する
    scrapejra.next_page_date(date_from.year, date_from.month)

    # 指定の日付の会場ごとのページにいく
    scrapejra.next_page_place()

    # レース詳細ページにいく
    scrapejra.next_page_place_race()

    # 馬ごとの情報を見る
    scrapejra.next_page_umadata()

    # Excel保存する
    exceldata.create_new_excel(date_from, date_to)


def get_data_updateexcel():
    # インスタンスを作る
    exceldata = ExcelData()
    scrapejra = ScrapeJRA()

    # Excelを読み込む
    if not exceldata.read_excel() :
        return

    # 出馬表を開く


if __name__ == '__main__':

    exceldata = ExcelData()
    scrapejra = ScrapeJRA()

    exceldata.read_excel()

    sel001 = input("1:過去データ取得　2:出馬表データ取得\n")
    # print(sel001)

    if sel001 == "1":
        print("新規作成")
        sel002 = input("データ取得開始日を入力してください。yyyymmdd\n")
        date_from = datetime.strptime(sel002, "%Y%m%d")
        # print(date_from)
        sel003 = input("データ取得終了日を入力してください。yyyymmdd\n")
        date_to = datetime.strptime(sel003, "%Y%m%d")
        # print(date_to)
        if date_to <= date_from:
            print("正しく入力してください。")
            exit()
        get_data_newexcel(date_from, date_to)
    elif sel001 == "2":
        # print("データ更新")
        get_data_updateexcel()
    else:
        print("正しく入力してください。")
        exit()

